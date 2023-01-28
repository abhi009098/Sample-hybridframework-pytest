import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl
path = "C:\\Users\\abhi0\\PycharmProjects\\hybridframework\\testdata\\Excel\\LoginData.xlsx"

@pytest.fixture
def driver():
    driver = webdriver.Chrome()
    driver.get("https://admin-demo.nopcommerce.com/")
    yield driver
    driver.quit()

def test_positive_login(driver):
    # Read test data from Excel
    workbook = openpyxl.load_workbook(path)
    sheet = workbook["Sheet1"]
    testdata = []
    for row in range(2, sheet.max_row + 1):
        data = {}
        data["username"] = sheet.cell(row, 1).value
        data["password"] = sheet.cell(row, 2).value
        data["exp"] = sheet.cell(row, 3).value
        testdata.append(data)

    for data in testdata:
        username = data["username"]
        password = data["password"]
        exp = data["exp"]

        # Find and fill in the username and password fields
        driver.find_element(By.CSS_SELECTOR, "input.email").clear()
        driver.find_element(By.CSS_SELECTOR, "input.email").send_keys(username)
        driver.find_element(By.CSS_SELECTOR, "input.password").clear()
        driver.find_element(By.CSS_SELECTOR, "input.password").send_keys(password)

        # Click the login button
        driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()

        if exp == "Pass":
            assert driver.find_element(By.XPATH, "(//a[normalize-space()='Logout'])[1]")
            driver.find_element(By.XPATH, "(//a[normalize-space()='Logout'])[1]").click()
        elif exp == "Fail":
            assert driver.find_element(By.CSS_SELECTOR, ".message-error")

