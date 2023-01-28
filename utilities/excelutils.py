import openpyxl
from openpyxl.styles import PatternFill


def get_row_count(file, sheet_name):
    """Get the number of rows in a worksheet."""
    # Load the workbook and get the worksheet
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    # Get the number of rows in the worksheet
    row_count = sheet.max_row
    return row_count


def get_column_count(file, sheet_name):
    """Get the number of columns in a worksheet."""
    # Load the workbook and get the worksheet
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    # Get the number of columns in the worksheet
    column_count = sheet.max_column
    return column_count


def read_cellData(file, sheet_name, row_no, column_no):
    """Read data from a cell in a worksheet and return it."""
    # Load the workbook and get the worksheet
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    # Get the cell at the specified row and column and get value
    return sheet.cell(row_no, column_no).value


def write_cellData(file, sheet_name, row_no, column_no, data):
    """Write data to a cell in a worksheet."""
    # Load the workbook and get the worksheet
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    # Get the cell at the specified row and column & set value or data of the cell
    sheet.cell(row_no, column_no).value = data
    # Save the workbook
    workbook.save(file)


def fill_green(file, sheet_name, row_no, column_no):
    """Fill a cell with a green background color."""
    # Load the workbook and get the worksheet
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    # Get the cell at the specified row and column
    cell = sheet.cell(row_no, column_no)
    # Create a green fill
    green_fill = PatternFill(fgColor='00FF00', fill_type='solid')
    # Set the fill of the cell
    cell.fill = green_fill
    # Save the workbook
    workbook.save(file)


def fill_red(file, sheet_name, row_no, column_no):
    """Fill a cell with a red background color."""
    # Load the workbook and get the worksheet
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    # Get the cell at the specified row and column
    cell = sheet.cell(row_no, column_no)
    # Create a red fill
    red_fill = PatternFill(fgColor='FF0000', fill_type='solid')
    # Set the fill of the cell
    cell.fill = red_fill
    # Save the workbook
    workbook.save(file)


from tabulate import tabulate


def read_worksheet_with_tabulate(file, sheet_name):
    """Read data from a worksheet in an Excel workbook using the tabulate module."""
    # Load the workbook and get the worksheet
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]

    # Read the data from the worksheet and store it in a list
    data = []
    for row in sheet.rows:
        row_data = []
        for cell in row:
            row_data.append(cell.value)
        data.append(row_data)

    # Convert the data to a tabular form using the tabulate() function
    table = tabulate(data)

    return table


# # table = read_worksheet_with_tabulate("example.xlsx", "Sheet1")
# print(table)


def write_worksheet_with_tabulate(file, sheet_name, data):
    """Write data to a worksheet in an Excel workbook using the tabulate module."""
    # Create a workbook and get the worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet(sheet_name)

    # Convert the data to a tabular form using the tabulate() function
    table = tabulate(data)

    # Split the table into rows and write each row to the worksheet
    for i, row in enumerate(table.split("\n")):
        for j, value in enumerate(row.split("\t")):
            cell = sheet.cell(row=i + 1, column=j + 1)
            cell.value = value

    # Save the workbook
    workbook.save(file)
